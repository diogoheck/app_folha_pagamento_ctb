from genericpath import exists
import os
from Ler_dados_folha import ler_dados
from Ler_dados_provisoes import ler_dados_prov_13, ler_dados_prov_ferias
from processamento_folha import folha as fopag
from processamento_provisao.provisoes import processar_provisao
from openpyxl import load_workbook
import os
import glob
from time import sleep


DADOS_LIDOS_FOLHA = './temp/dados_lidos_folha.txt'
empresas = './banco_de_dados/empresas.xlsx'
centro_custos = './banco_de_dados/centro_de_custos.xlsx'
eventos = './banco_de_dados/eventos.xlsx'
arquivo_folha = './relatorios/RELATORIO_ESPELHO_RESUMO.xlsx'
arquivo_folha_13_1 = './relatorios/Relatório de provisão de 13º salário_1.xlsx'
arquivo_folha_13_2 = './relatorios/Relatório de provisão de 13º salário_2.xlsx'
arquivo_folha_ferias_1 = './relatorios/Relatório de provisão de férias_1.xlsx'
arquivo_folha_ferias_2 = './relatorios/Relatório de provisão de férias_2.xlsx'
DATA_ARQUIVO_FOLHA_FERIAS = './temp/data_arquivo_prov_ferias.txt'
DATA_ARQUIVO_FOLHA_13 = './temp/data_arquivo_prov_13.txt'


def remover_arquivo(arquivo):
    if os.path.exists(arquivo):
        os.remove(arquivo)


def remover_todos(caminho):
    py_files = glob.glob(caminho)

    for py_file in py_files:
        try:
            os.remove(py_file)
        except OSError as e:
            print(f"Error:{ e.strerror}")


def montar_dicionarios(planilha, indicador=False, indice=0):
    dicionario_planilha = {}

    for linha_planilha in planilha.values:
        if indicador:
            dicionario_planilha[str(linha_planilha[indice])] = linha_planilha
        else:
            dicionario_planilha[str(linha_planilha[indice])] = linha_planilha

    return dicionario_planilha


if __name__ == '__main__':

    dic_nome_centro_custos = ''
    lista_sistemas = ['1', '2', '3', '4', '5', '6', '7', '8']
    sistema = '0'
    adiantamento = -1
    while sistema not in lista_sistemas:
        print('')
        print('===========================================')
        print('Bem vindo ao Sistema de Importação da Folha')
        print('===========================================')
        print('')
        print('*** Escolha o número desejado de acordo com a empresa: ***')
        print('')
        print('1 - Allianza-Imex')
        print('2 - MPE')
        print('3 - DELLA-PASQUA')
        print('4 - Multi')
        print('5 - SB-Rubenich')
        print('6 - Camsul')
        print('7 - Super Safra')
        print('8 - Cootranscau')
        print('')
        sistema = input()
        

    if sistema == '1':
        sistema = 'Allianza-Imex'
    elif sistema == '2':
        sistema = 'MPE'
    elif sistema == '3':
        sistema = 'DELLA-PASQUA'
    elif sistema == '4':
        sistema = 'MULTI'
    elif sistema == '5':
        sistema = 'SB-Rubenich'
    elif sistema == '6':
        sistema = 'Camsul'
    elif sistema == '7':
        sistema = 'Super-Safra'
    elif sistema == '8':
        sistema = 'Cootranscau'

    print('\n')
    print(f"Você selecionou a empresa  ==>  *** {sistema} ***")
    print('\n')

    print('processando folha......')
    print('\n')
    sleep(2)
    # while adiantamento != 'a' and adiantamento != 'f':
    #     print('')
    #     print('Digite a "f" para folha normal ou "a" para folha adiantamento')
    #     adiantamento = input()
        # print(adiantamento)

    
    print('removendo arquivos temporarios e de folha e provisoes........')
    print('\n')
    sleep(2)
    # remover_arquivo('./arquivo_gerado/arquivo_importacao_folha.txt')
    remover_arquivo('./temp/data_arquivo_folha.txt')
    remover_arquivo('./temp/data_arquivo_prov.txt')
    remover_arquivo('./temp/data_arquivo_prov_ferias.txt')
    remover_todos('./arquivo_gerado/*.*')
    remover_todos('./temp/*.*')

    print('carregando os relatórios de folha e provisões....')
    print('\n')
    sleep(2)

    if exists(arquivo_folha_13_1) and exists(arquivo_folha_13_2):
        # carregar arquivo xlsx folha pagto
        folha_13 = './relatorios/Relatório de provisão de 13º salário_1.xlsx'
        folha_13 = load_workbook(folha_13)
        plan_prov_13 = folha_13.active

        # carregar arquivo xlsx folha pagto
        folha_13 = './relatorios/Relatório de provisão de 13º salário_2.xlsx'
        folha_13 = load_workbook(folha_13)
        plan_prov_13_pos = folha_13.active

    if exists(arquivo_folha_ferias_1) and exists(arquivo_folha_ferias_2):
        # carregar arquivo xlsx folha pagto
        folha_ferias = './relatorios/Relatório de provisão de férias_1.xlsx'
        folha_ferias = load_workbook(folha_ferias)
        plan_prov_ferias = folha_ferias.active

        # carregar arquivo xlsx folha pagto
        folha_ferias = './relatorios/Relatório de provisão de férias_2.xlsx'
        folha_ferias = load_workbook(folha_ferias)
        plan_prov_ferias_pos = folha_ferias.active

    # if adiantamento == 'f':
    #     # carregar layout da folha
    #     if sistema == 'DELLA-PASQUA':
    #         layout_folha = './layouts/layout_folha_della_pasqua.xlsx'
    #     else:
    #         layout_folha = './layouts/layout_folha_geral.xlsx'

    #     layout_folha = load_workbook(layout_folha)
    #     layout_folha_pagto = layout_folha.active
    # else:
    #     # carregar layout da folha
    #     if sistema == 'DELLA-PASQUA':
    #         layout_folha = './layouts/layout_adiantamento_folha.xlsx'
    #     else:
    #         layout_folha = './layouts/layout_adiantamento_folha.xlsx'

    #     layout_folha = load_workbook(layout_folha)
    #     layout_folha_pagto = layout_folha.active

    if exists(arquivo_folha):
        # carregar arquivo xlsx folha pagto
        try:
            folha = './relatorios/RELATORIO_ESPELHO_RESUMO.xlsx'
            folha = load_workbook(folha)
            plan_folha = folha.active
            # folha.save('./relatorios/RELATORIO_ESPELHO_RESUMO.xlsx')
        except PermissionError as e:
            print('RELATORIO_ESPELHO_RESUMO.xlsx está aberto')
            print('\n')
            os.system('pause')
            exit(1)
    print('carregando os bancos de dados da empresa.....')
    print('\n')
    sleep(3)

    if sistema == 'DELLA-PASQUA':
        empresas = './banco_de_dados/empresas_della.xlsx'
        centro_custos = './banco_de_dados/centro_de_custos_della.xlsx'
        eventos = './banco_de_dados/eventos_della.xlsx'
    elif sistema == 'MULTI':
        empresas = './banco_de_dados/empresas_multi.xlsx'
        centro_custos = './banco_de_dados/centro_de_custos_multi.xlsx'
        eventos = './banco_de_dados/eventos_multi.xlsx'
    elif sistema == 'MPE':
        empresas = './banco_de_dados/empresas_MPE.xlsx'
        centro_custos = './banco_de_dados/centro_de_custos_MPE.xlsx'
        eventos = './banco_de_dados/eventos_MPE.xlsx'
    elif sistema == 'Allianza-Imex':
        empresas = './banco_de_dados/empresas_Allianza_Imex.xlsx'
        centro_custos = './banco_de_dados/centro_de_custos_Allianza_Imex.xlsx'
        eventos = './banco_de_dados/eventos_Allianza_Imex.xlsx'
    elif sistema == 'SB-Rubenich':
        empresas = './banco_de_dados/empresas_SB.xlsx'
        centro_custos = './banco_de_dados/centro_de_custos_SB.xlsx'
        eventos = './banco_de_dados/eventos_SB.xlsx'
    elif sistema == 'Camsul':
        empresas = './banco_de_dados/empresas_Camsul.xlsx'
        eventos = './banco_de_dados/eventos_Camsul.xlsx'
    elif sistema == 'Cootranscau':
        empresas = './banco_de_dados/empresas_Cootranscau.xlsx'
        eventos = './banco_de_dados/eventos_Cootranscau.xlsx'
        centro_custos = './banco_de_dados/centro_de_custos_cootranscau.xlsx'
    elif sistema == 'Super-Safra':
        empresas = './banco_de_dados/empresas_super_safra.xlsx'
        eventos = './banco_de_dados/eventos_super_safra.xlsx'
        centro_custos = './banco_de_dados/centro_de_custos_super_safra.xlsx'
    
        

    if exists(empresas):
        empresas = load_workbook(empresas)
        plan_empresas = empresas.active
        dic_nome_empresas = montar_dicionarios(plan_empresas)
    if exists(centro_custos):
        centro_custos = load_workbook(centro_custos)
        plan_centro_custos = centro_custos.active
        dic_nome_centro_custos = montar_dicionarios(plan_centro_custos)
    if exists(eventos):
        eventos = load_workbook(eventos)
        plan_eventos = eventos.active
        dic_eventos = montar_dicionarios(plan_eventos)

    # print(dic_eventos)

    if exists(arquivo_folha):
        # ler dados folha
        print('lendo os dados da folha......')
        print('\n')
        sleep(2)
        dados_lidos_folha = ler_dados.ler_folha_completa(plan_folha)
        # print(dados_lidos_folha)
        # gerar arquivo folha
        print('processando o arquivo para importacao folha.....')
        print('\n')
        sleep(2)
        fopag.processar_folha(dados_lidos_folha, dic_nome_empresas,
                              dic_nome_centro_custos, dic_eventos, sistema)

    if exists(arquivo_folha_13_1) and exists(arquivo_folha_13_2):
        print('lendo os dados das provisoes de 13 salario......')
        print('\n')
        sleep(2)
        dados_lidos_prov_13_1 = ler_dados_prov_13.ler_folha_provisao_13(
            plan_prov_13)
        dados_lidos_prov_13_2 = ler_dados_prov_13.ler_folha_provisao_13(
            plan_prov_13_pos)
        print('gerando o arquivo de importacao da provisao do 13 salario....')
        print('\n')
        sleep(2)
        processar_provisao(dados_lidos_prov_13_1, dados_lidos_prov_13_2, dic_nome_empresas,
                           dic_nome_centro_custos, dic_eventos, sistema, DATA_ARQUIVO_FOLHA_13, DECIMO=True)

    if exists(arquivo_folha_ferias_1) and exists(arquivo_folha_ferias_2):
        print('lendo os dados das provisoes de ferias......')
        print('\n')
        sleep(2)
        dados_lidos_prov_ferias_1 = ler_dados_prov_ferias.ler_folha_provisao_ferias(
            plan_prov_ferias)
        dados_lidos_prov_ferias_2 = ler_dados_prov_ferias.ler_folha_provisao_ferias(
            plan_prov_ferias_pos)
        print('gerando o arquivo de importacao da provisao de ferias....')
        print('\n')
        sleep(2)
        processar_provisao(dados_lidos_prov_ferias_1, dados_lidos_prov_ferias_2,
                           dic_nome_empresas,
                           dic_nome_centro_custos, dic_eventos, sistema, DATA_ARQUIVO_FOLHA_FERIAS, DECIMO=False)

    print('removendo arquivos temporarios')
    print('\n')
    sleep(2)
    remover_todos('./temp/*.*')

    print('\n\n**********  processo finalizado!!!!!! \o/\o/ **********\n\n')
    os.system("pause")
